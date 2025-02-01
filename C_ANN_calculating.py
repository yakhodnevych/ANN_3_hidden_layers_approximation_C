import openpyxl
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from os.path import join, abspath

#обчислення значення досліджуваної величини С (коеф. Шезі)
# за допомогою навченої ШНМ з 2 прихованими шарами (налаштованих матриць вагових коефіцієнтів),
# кількість нейронів однакова для всіх прихованих шарів,
# матриці ваг завантажуються з текстових файлів,
# вхідні аргументи для шуканої функції завантажуються з файлу ексель,
# результати обчислень зберігаються у файлі ексель

# визначення функції для обчислення стандартної логістичної ф-ії активації
betta = 1  #параметр логістичної функції, за замовчуванням = 1, впливає на швидкість навчання

def logistic(x):
    return 1.0 / (1 + np.exp(-x*betta))

#зчитування матриць вагових коефіцієнтів
def OpenFileToMatrix (f_name):
    data_path = join('.', 'Data', f_name)
    data_path = abspath(data_path)
    f = open(data_path)
    raw_matrix = f.readlines()
    f.close()
    return raw_matrix

# відкриваємо вказаний файл тільки для читання значень (формули не зчитуємо)
data_path = join('.',"Input.xlsx") #'.' - шлях до поточної директорії (де запущено програму) включно, об*єднується з назвою файлу, який будемо відкривати
data_path = abspath(data_path) #створюється повний шлях для відкриття файлу

#відкриваємо файл у поточній директорії (там де запущено програму)
Input_Data = openpyxl.open(data_path, read_only=True, data_only=True)
sheet_data = Input_Data.worksheets[0]

max_row_data = sheet_data['A3'].value  # кількість прикладів вхідних даних, для яких будемо обчислювати шукану величину

# параметри мережі, що відповідають вхідним даним
input_size = sheet_data['A5'].value  # кількість входів
hidden_size = sheet_data['A7'].value  # к-сть нейронів у прихованому шарі (однакова для всіх прихованих шарів)
output_size = sheet_data['A9'].value  # кількість виходів
hidden_layers_c = 3  # к-сть прихованих шарiв

# визначення та ініціалізація матриці вхідних значень і матриці обчислених значень С

len_ryadok = input_size

characteristics_riverbed = np.zeros(
    shape=(max_row_data, len_ryadok))  # вхід - гідродинамічні характеристики русла

coef_C = np.zeros(shape=(max_row_data))  # вихід - значення коефіцієнта Шезі

# завантаження вхідних значень
for i in range(max_row_data):
    for j in range(len_ryadok):
        characteristics_riverbed[i][j] = sheet_data[i + 2][j + 1].value

print('Завантажено вхiднi данi для штучної нейронної мережi з файлу:')
print(data_path)
print('Параметри мережi:')
print('кiлькiсть входiв -', input_size)
print('к-сть прихованих шарiв:', hidden_layers_c)
print('к-сть нейронiв у прихованому шарi -', hidden_size)
print('кiлькiсть виходiв -', output_size)
print('Кiлькiсть вхiдних прикладiв:', max_row_data)

list_param=[] # список описів вхідних параметрів
print('Дослiджується коеф. С/100 з врахуванням параметрiв', end=' ')
for j in range(len_ryadok):
    list_param.append(sheet_data[1][j+1].value)
    print(sheet_data[1][j+1].value, end=',')
print()

Input_Data.close() # закриття файлу вхідних даних Excel

#зчитування матриць вагових коефіцієнтів W_1, W_1_ab і W_2
raw_matrix1 = OpenFileToMatrix ("weights_matrix_1.txt")
raw_matrix1_ab = OpenFileToMatrix ("weights_matrix_1_ab.txt")
raw_matrix1_bc = OpenFileToMatrix ("weights_matrix_1_bc.txt")
raw_matrix2 = OpenFileToMatrix ("weights_matrix_2.txt")

#ініціалізація матриць W_1, W_1_ab, W_1_bc і W_2
W_1 = np.zeros(
    shape=(input_size, hidden_size))
W_1_ab = np.zeros(
    shape=(hidden_size, hidden_size))
W_1_bc = np.zeros(
    shape=(hidden_size, hidden_size))
W_2 = np.zeros(shape=(hidden_size, output_size))

#заповненення матриць W_1, W_1_ab і W_2

for i in range (len(raw_matrix1)):
    b=raw_matrix1[i]
    a=b.split(' ')
    # len(a) == len(raw_matrix2) - к-сть рядків матриці W_2 = к-сті стовбців W_1
    for j in range(len(raw_matrix2)):
        W_1[i][j]=float(a[j])

for i in range (len(raw_matrix1_ab)):
    b=raw_matrix1_ab[i]
    a=b.split(' ')
    # len(a) == len(raw_matrix2) - к-сть рядків матриці W_2 = к-сті стовбців W_1
    for j in range(len(raw_matrix2)):
        W_1_ab[i][j]=float(a[j])

for i in range (len(raw_matrix1_bc)):
    b=raw_matrix1_bc[i]
    a=b.split(' ')
    # len(a) == len(raw_matrix2) - к-сть рядків матриці W_2 = к-сті стовбців W_1
    for j in range(len(raw_matrix2)):
        W_1_bc[i][j]=float(a[j])

for i in range (len(raw_matrix2)):
    W_2[i]=float(raw_matrix2[i])

if (len(raw_matrix1)==input_size)and(len(raw_matrix2)==hidden_size):
    print('вхiднi данi вiдповiдають параметрам матриць вагових коефiцiєнтiв ШНМ,')
    # обчислення виходів мережі
    for i in range(max_row_data):
        # прямий хід обчислень
        layer_0 = characteristics_riverbed[i:i + 1]
        layer_1a = logistic(np.dot(layer_0, W_1))
        layer_1b = logistic(np.dot(layer_1a, W_1_ab))
        layer_1c = logistic(np.dot(layer_1b, W_1_bc))
        layer_2 = np.dot(layer_1c, W_2)
        coef_C[i] = layer_2
        print('для набору параметрiв №',i+1,', обчислене C/100 =', coef_C[i])


    Out_Data=Workbook()
    sheet_data=Out_Data.active
    sheet_data.title="output_data"
    # запис даних в таблицю

    # додаються назви стовбців
    sheet_data['A1'].value='j'
    sheet_data.cell(row=1, column=len_ryadok+2).value='C/100'

    for j in range(len_ryadok):
        sheet_data[1][j + 1].value=list_param[j]

    # записуємо вхідні дані та відповідні обчислені значення С
    for i in range(max_row_data):
        sheet_data['A'+str(i+2)].value = i+1

        sheet_data.cell(row=i+2, column=len_ryadok + 2).value = coef_C[i]
        for j in range(len_ryadok):
            sheet_data[i + 2][j + 1].value = characteristics_riverbed[i][j]


    #форматування
    for i in range(1, len_ryadok+2):
        zag = sheet_data.cell(row=1, column=i)
        zag.alignment = Alignment(horizontal='center')
        zag.font = Font(bold=True, italic=False, color='DC143C', size=12)

        zag = sheet_data.cell(row=1, column=len_ryadok + 2)
        zag.alignment = Alignment(horizontal='center')
        zag.font = Font(bold=True, italic=False, color='000000', size=12)

    for i in range(max_row_data+1):
        sheet_data.cell(row=i+1, column=len_ryadok+2).fill=PatternFill(
            fill_type='solid', start_color='90EE90', end_color='90EE90'
        )


    exfilename = join('.', ('Output.xlsx')) # створення (об'єднання з частин)  імені файла для зберігання
    exfilename=abspath(exfilename) #створення повного шляху, де буде зберігатись файл

    Out_Data.save(exfilename)  # зберігання файлу
    Out_Data.close()
    print("Результати обчислення збереженi у файлi: ")
    print(exfilename)

else:
    print('Вхiднi данi не відповiдають параметрам матриць ваг ШНМ.')
    print('Перевiрте вiдповiднiсть кiлькостi входiв та нейронiв прихованого шару')
    print(' у файлi вхiдних даних параметрам навченої мережi:')
    print('кiлькiсть входiв мережi =', len(raw_matrix1))
    print('кiлькiсть прихованих шарiв мережi =', hidden_layers_c)
    print('кiлькiсть  нейронiв у прихованому шарi мережi =', len(raw_matrix2))

input('Натиснiть будь-яку клавiшу для завершення програми.')